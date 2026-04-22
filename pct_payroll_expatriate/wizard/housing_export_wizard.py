from odoo import fields, models
from odoo.exceptions import UserError
import base64
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


class HousingExportWizard(models.TransientModel):
    _name = 'housing.export.wizard'
    _description = 'Housing Export Wizard'

    housing_ids = fields.Many2many(
        'expatriate.housing',
        string='Housing Records'
    )
    excel_file = fields.Binary(
        string='Excel File',
        readonly=True
    )
    filename = fields.Char(
        string='Filename',
        default='housing_report.xlsx'
    )

    def _get_cost_category(self, cost_line, rent_product, maintenance_product, electricity_product):
        """Categorize cost line into rent, maintenance, or electricity"""
        if cost_line.product_id:
            if rent_product and cost_line.product_id.id == rent_product.id:
                return 'rent'
            elif maintenance_product and cost_line.product_id.id == maintenance_product.id:
                return 'maintenance'
            elif electricity_product and cost_line.product_id.id == electricity_product.id:
                return 'electricity'
            else:
                # Fallback: categorize by product name
                product_name = (cost_line.product_id.name or '').lower()
                if 'rent' in product_name:
                    return 'rent'
                elif 'maintenance' in product_name or 'diesel' in product_name:
                    return 'maintenance'
                elif 'electric' in product_name or 'nepa' in product_name:
                    return 'electricity'
        return 'other'

    def _calculate_employee_costs(self, housing, employee, rent_product, maintenance_product, electricity_product):
        """Calculate costs for a specific employee in a housing unit"""
        employee_costs = {'rent': 0, 'maintenance': 0, 'electricity': 0}
        total_employees = len(housing.employee_ids)

        for cost_line in housing.cost_line_ids:
            category = self._get_cost_category(cost_line, rent_product, maintenance_product, electricity_product)
            if category == 'other':
                continue

            if cost_line.employee_ids:
                # Cost is assigned to specific employees
                if employee in cost_line.employee_ids:
                    # Split among assigned employees
                    share = cost_line.amount / len(cost_line.employee_ids)
                    employee_costs[category] += share
            else:
                # Cost is shared among all housing occupants
                if total_employees > 0:
                    share = cost_line.amount / total_employees
                    employee_costs[category] += share

        return employee_costs

    def action_export(self):
        """Generate Excel report with housing summary and employee breakdown"""
        if not Workbook:
            raise UserError('Python library openpyxl is not installed. Please install it to export Excel files.')

        # Get housing records
        housings = self.housing_ids
        if not housings:
            raise UserError('No housing records to export.')

        # Create workbook
        wb = Workbook()

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get product references for cost categorization
        rent_product = self.env.ref('pct_payroll_expatriate.product_housing_rent', raise_if_not_found=False)
        maintenance_product = self.env.ref('pct_payroll_expatriate.product_housing_maintenance', raise_if_not_found=False)
        electricity_product = self.env.ref('pct_payroll_expatriate.product_housing_electricity', raise_if_not_found=False)

        # ==================== SHEET 1: Housing Summary ====================
        ws_housing = wb.active
        ws_housing.title = "Housing Summary"

        # Write headers
        housing_headers = ['Housing', 'Employees', '# Occupants', 'Location', 'Type', 'Contract Start', 'Contract End', 'Days to Expire', 'Alert Status', 'Rent', 'Maintenance', 'Electricity', 'Total Cost']
        for col_num, header in enumerate(housing_headers, 1):
            cell = ws_housing.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write housing data
        row_num = 2
        total_rent = 0
        total_maintenance = 0
        total_electricity = 0
        grand_total = 0

        for housing in housings:
            # Get cost breakdown by product
            rent = 0
            maintenance = 0
            electricity = 0

            for cost_line in housing.cost_line_ids:
                category = self._get_cost_category(cost_line, rent_product, maintenance_product, electricity_product)
                if category == 'rent':
                    rent += cost_line.amount
                elif category == 'maintenance':
                    maintenance += cost_line.amount
                elif category == 'electricity':
                    electricity += cost_line.amount

            row_total = rent + maintenance + electricity
            total_rent += rent
            total_maintenance += maintenance
            total_electricity += electricity
            grand_total += row_total

            # Get alert status display value
            alert_status_display = ''
            if housing.alert_status:
                alert_status_display = dict(housing._fields['alert_status'].selection).get(housing.alert_status, '')

            # Get employee names
            employee_names = ', '.join(housing.employee_ids.mapped('name'))

            row_data = [
                housing.name or '',
                employee_names,
                len(housing.employee_ids),
                housing.location or '',
                housing.housing_type or '',
                housing.contract_start_date.strftime('%d-%b-%Y') if housing.contract_start_date else '',
                housing.contract_end_date.strftime('%d-%b-%Y') if housing.contract_end_date else '',
                housing.days_to_expire,
                alert_status_display,
                rent,
                maintenance,
                electricity,
                row_total
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws_housing.cell(row=row_num, column=col_num, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

            row_num += 1

        # Write totals row
        total_row = row_num
        ws_housing.cell(row=total_row, column=1, value='TOTALS').font = Font(bold=True)
        ws_housing.cell(row=total_row, column=1).border = thin_border

        for col in range(2, 10):
            ws_housing.cell(row=total_row, column=col, value='').border = thin_border

        totals = [total_rent, total_maintenance, total_electricity, grand_total]
        for col_num, total_value in enumerate(totals, 10):
            cell = ws_housing.cell(row=total_row, column=col_num, value=total_value)
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Adjust column widths
        housing_widths = [30, 35, 12, 25, 15, 15, 15, 15, 12, 15, 15, 15, 15]
        for i, width in enumerate(housing_widths, 1):
            ws_housing.column_dimensions[get_column_letter(i)].width = width

        ws_housing.freeze_panes = 'A2'

        # ==================== SHEET 2: Employee Cost Breakdown ====================
        ws_employee = wb.create_sheet(title="Employee Costs")

        # Write headers
        employee_headers = ['Employee', 'Job Description', 'Site Location', 'Housing', 'Location', 'Rent', 'Maintenance', 'Electricity', 'Total Cost']
        for col_num, header in enumerate(employee_headers, 1):
            cell = ws_employee.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write employee data
        row_num = 2
        emp_total_rent = 0
        emp_total_maintenance = 0
        emp_total_electricity = 0
        emp_grand_total = 0

        for housing in housings:
            for employee in housing.employee_ids:
                costs = self._calculate_employee_costs(housing, employee, rent_product, maintenance_product, electricity_product)

                employee_total = costs['rent'] + costs['maintenance'] + costs['electricity']
                emp_total_rent += costs['rent']
                emp_total_maintenance += costs['maintenance']
                emp_total_electricity += costs['electricity']
                emp_grand_total += employee_total

                row_data = [
                    employee.name or '',
                    employee.job_id.name if employee.job_id else '',
                    employee.work_location_id.name if employee.work_location_id else '',
                    housing.name or '',
                    housing.location or '',
                    round(costs['rent'], 2),
                    round(costs['maintenance'], 2),
                    round(costs['electricity'], 2),
                    round(employee_total, 2)
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws_employee.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

                row_num += 1

        # Write totals row
        total_row = row_num
        ws_employee.cell(row=total_row, column=1, value='TOTALS').font = Font(bold=True)
        ws_employee.cell(row=total_row, column=1).border = thin_border

        for col in range(2, 6):
            ws_employee.cell(row=total_row, column=col, value='').border = thin_border

        emp_totals = [round(emp_total_rent, 2), round(emp_total_maintenance, 2), round(emp_total_electricity, 2), round(emp_grand_total, 2)]
        for col_num, total_value in enumerate(emp_totals, 6):
            cell = ws_employee.cell(row=total_row, column=col_num, value=total_value)
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Adjust column widths
        employee_widths = [25, 20, 20, 30, 25, 15, 15, 15, 15]
        for i, width in enumerate(employee_widths, 1):
            ws_employee.column_dimensions[get_column_letter(i)].width = width

        ws_employee.freeze_panes = 'A2'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        # Update wizard with file
        self.excel_file = base64.b64encode(output.read())
        self.filename = f'housing_report_{fields.Date.today()}.xlsx'
        output.close()

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'housing.export.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
